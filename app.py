import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Steam RRP Generator", layout="wide")
st.title("🎮 Steam RRP Price Generator")

@st.cache_data
def load_ssrp(csv_file):
    df_raw = pd.read_csv(csv_file, sep=';', header=None)
    country_codes = df_raw.iloc[0].tolist()
    currencies = df_raw.iloc[1].tolist()
    data = df_raw.iloc[2:].reset_index(drop=True)
    data.columns = country_codes
    return country_codes, currencies, data

def fetch_steam_price(app_id):
    url = f"https://store.steampowered.com/api/appdetails?appids={app_id}&cc=us&l=en"
    r = requests.get(url)
    try:
        json_data = r.json()
        price_data = json_data[str(app_id)]['data']['price_overview']
        price = price_data['initial'] / 100  # Steam stores price in cents
        currency = price_data['currency']
        return price, currency
    except:
        return None, None

def calculate_prices(base_price_usd, ssrp_df, vat_map):
    results = []
    for country in ssrp_df.columns:
        try:
            row = ssrp_df[country].dropna().astype(float)
            recommended_price = row.values[0]
            vat = vat_map.get(country.lower(), 0)
            adjusted_price = round(recommended_price * (1 + vat / 100), 2)
            results.append({
                "Country": country.upper(),
                "Steam RRP": recommended_price,
                "VAT %": vat,
                "Final Price": adjusted_price
            })
        except:
            continue
    return pd.DataFrame(results)

# --- UI ---
app_id = st.text_input("Enter Steam App ID")

uploaded_file = st.file_uploader("Upload SSRP.csv file", type=["csv"])

vat_rates = {
    'us': 0, 'eu': 21, 'gb': 20, 'ru': 20, 'cn': 13, 'au': 10, 'br': 17, 'ca': 5,
    'jp': 10, 'mx': 16, 'kr': 10, 'tr': 18, 'za': 15, 'ch': 8, 'in': 18
}

if app_id and uploaded_file:
    st.write("📥 Fetching base price from Steam...")
    base_price, base_currency = fetch_steam_price(app_id)

    if base_price:
        st.success(f"Base Steam Price: {base_price} {base_currency}")
        st.write("📊 Calculating regional prices from SSRP...")
        codes, currencies, ssrp_data = load_ssrp(uploaded_file)

        results_df = calculate_prices(base_price, ssrp_data, vat_rates)
        st.dataframe(results_df)

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Steam RRP Results"
        ws.append(results_df.columns.tolist())
        for row in results_df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        wb.save(output)

        st.download_button(
            label="📥 Download Excel",
            data=output.getvalue(),
            file_name="steam_rrp_prices.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("❌ Could not fetch price from Steam. Check App ID.")
